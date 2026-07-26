from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import HTTPException, Request, Response, Depends, Header, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from typing import List, Optional, Dict, Any
from bson import ObjectId
from datetime import datetime, timezone, timedelta
import logging, uuid, jwt, bcrypt, io, json, requests, random, stripe, httpx, hashlib, secrets
from emergentintegrations.llm.chat import LlmChat, UserMessage, TextDelta, StreamDone

# ---------------------------------------------------------------- config
JWT_ALGORITHM = "HS256"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
APP_NAME = "ceo-ai"
stripe.api_key = os.environ.get("STRIPE_SECRET_KEY") or "sk_test_emergent"
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET", "")
EMAIL_BASE_URL = "https://integrations.emergentagent.com"
EMAIL_KEY = os.environ.get("EMERGENT_EMAIL_KEY")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "CEO AI")

MODEL_MAP = {
    "claude": ("anthropic", "claude-opus-4-7"),
    "gpt": ("openai", "gpt-5.5"),
    "gemini": ("gemini", "gemini-3.1-pro-preview"),
}
CURRENCY_SYMBOL = {"EUR": "€", "BRL": "R$", "USD": "$"}

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------- auth helpers
def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True, secure=True,
                        samesite="none", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizador não encontrado")
        user["id"] = str(user["_id"])
        user["is_premium"] = bool(user.get("is_premium"))
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sessão expirada")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def is_premium(user_id: str) -> bool:
    u = await db.users.find_one({"_id": ObjectId(user_id)})
    return bool(u and u.get("is_premium"))

# ---------------------------------------------------------------- company resolution
async def resolve_company(user_id: str, company_id: Optional[str] = None):
    if company_id:
        c = await db.companies.find_one({"_id": ObjectId(company_id), "user_id": user_id})
        if c:
            return c
    s = await db.settings.find_one({"user_id": user_id}) or {}
    acid = s.get("active_company_id")
    if acid:
        c = await db.companies.find_one({"_id": ObjectId(acid), "user_id": user_id})
        if c:
            return c
    return await db.companies.find_one({"user_id": user_id})

async def active_company_id(user_id: str) -> Optional[str]:
    c = await resolve_company(user_id)
    if not c:
        return None
    cid = str(c["_id"])
    # migrate orphan entries to the active company (one-time, cheap)
    await db.entries.update_many({"user_id": user_id, "company_id": {"$exists": False}},
                                 {"$set": {"company_id": cid}})
    return cid

# ---------------------------------------------------------------- storage
storage_key = None
def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key, "Content-Type": content_type}, data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()

def extract_document_text(data: bytes, content_type: str, filename: str) -> str:
    name = (filename or "").lower(); ct = (content_type or "").lower()
    try:
        if name.endswith(".pdf") or "pdf" in ct:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            return "\n".join((p.extract_text() or "") for p in reader.pages[:25])
        if name.endswith(".xlsx") or "sheet" in ct or "excel" in ct:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets[:6]:
                out.append(f"# {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        out.append("\t".join(cells))
                    if len(out) > 2500:
                        break
            return "\n".join(out)
        if name.endswith(".docx") or "wordprocessing" in ct:
            import docx
            d = docx.Document(io.BytesIO(data))
            return "\n".join(p.text for p in d.paragraphs)
        if name.endswith((".csv", ".txt", ".tsv")) or ct.startswith("text"):
            return data.decode("utf-8", errors="ignore")
    except Exception as e:
        logger.error(f"extract_document_text error: {e}")
    return ""

async def analyze_document(text: str, doc_type: str, filename: str) -> dict:
    if not text or len(text.strip()) < 20:
        return {"analysable": False, "relevant": False, "quality": "low",
                "summary": "Carregado, mas sem texto legível para análise automática (ex: imagem ou digitalização)."}
    prompt = (
        f"És um analista financeiro. Analisa este documento ('{filename}', categoria '{doc_type}') e devolve APENAS JSON: "
        '{"analysable":true,"relevant":bool,"quality":"high"|"medium"|"low","doc_kind":str,"period":str,"summary":str,'
        '"figures":{"revenue":number|null,"ebitda":number|null,"net_profit":number|null,"assets":number|null,'
        '"liabilities":number|null,"recurring_revenue":number|null,"currency":str|null},'
        '"contracts":{"count":number|null,"total_value":number|null,"recurring":bool}|null,'
        '"red_flags":[str],"strengths":[str]}. '
        "'relevant'=true se contém dados financeiros/contratuais úteis para avaliar a empresa. "
        "'quality': 'high' se são demonstrações/contratos formais com números claros; 'medium' se parcial; 'low' se pouco útil. "
        "Preenche 'figures' APENAS com números presentes no documento (senão null). 'summary' em 1-2 frases. "
        "Português europeu. Sem texto fora do JSON.\n\nCONTEÚDO:\n" + text[:8000]
    )
    ai = await ai_json("És um analista financeiro rigoroso. Respondes só com JSON.", prompt)
    return ai or {"analysable": True, "relevant": False, "quality": "low", "summary": "Não foi possível analisar o documento."}


def rag(value, good, warn, reverse=False):
    if reverse:
        if value <= good: return "green"
        if value <= warn: return "amber"
        return "red"
    if value >= good: return "green"
    if value >= warn: return "amber"
    return "red"

def compute_balance(company: dict, profile: dict, entries_net: float = 0.0):
    """Single source of truth for the company balance sheet."""
    profile = profile or {}
    has = bool(profile)
    cash = float(profile.get("cash_balance", 0) or 0) if has else (float((company or {}).get("bank_balance", 0) or 0) + entries_net)
    assets_items = sum(float(a.get("amount", 0) or 0) for a in (profile.get("assets") or []))
    liab_items = sum(float(l.get("amount", 0) or 0) for l in (profile.get("liabilities") or []))
    debt = float(profile.get("total_debt", 0) or 0)
    total_assets = cash + assets_items
    total_liabilities = debt + liab_items
    return {"cash": round(cash, 2), "total_assets": round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "net_worth": round(total_assets - total_liabilities, 2)}

async def build_snapshot(user_id: str):
    company = await resolve_company(user_id) or {}
    cid = str(company["_id"]) if company.get("_id") else None
    entries = await db.entries.find({"user_id": user_id, "company_id": cid}, {"type": 1, "amount": 1, "date": 1, "category": 1}).to_list(5000) if cid else []
    now = datetime.now(timezone.utc)
    month_key = now.strftime("%Y-%m")
    income = sum(e["amount"] for e in entries if e["type"] == "income")
    expense = sum(e["amount"] for e in entries if e["type"] == "expense")
    m_income = sum(e["amount"] for e in entries if e["type"] == "income" and str(e.get("date", "")).startswith(month_key))
    m_expense = sum(e["amount"] for e in entries if e["type"] == "expense" and str(e.get("date", "")).startswith(month_key))
    net = income - expense
    m_net = m_income - m_expense
    profile = await db.financial_profiles.find_one({"user_id": user_id, "company_id": cid}) if cid else None
    bal = compute_balance(company, profile, net)
    bank = bal["cash"]
    monthly_burn = m_expense if m_expense > 0 else (expense / 12 if expense else 1)
    runway = bank / monthly_burn if monthly_burn > 0 else 99
    profit_margin = (m_net / m_income * 100) if m_income > 0 else 0
    tax_reserve = float(company.get("monthly_tax_estimate", 0))
    payroll = sum(e["amount"] for e in entries if e["type"] == "expense" and "salári" in str(e.get("category", "")).lower())
    currency = company.get("currency", "EUR")

    vitals = [
        {"key": "cashflow", "label": "Fluxo de Caixa", "value": round(m_net, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(m_net, 0.01, -monthly_burn * 0.5), "hint": "Entradas menos saídas este mês"},
        {"key": "profit", "label": "Lucro", "value": round(profit_margin, 1), "unit": "%",
         "status": rag(profit_margin, 10, 3), "hint": "Margem de lucro mensal"},
        {"key": "clients", "label": "Clientes", "value": int(company.get("clients_count", 0)), "unit": "",
         "status": rag(company.get("clients_count", 0), 10, 3), "hint": "Base de clientes ativa"},
        {"key": "tax", "label": "Impostos", "value": round(tax_reserve, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(bank - tax_reserve, 0.01, -1), "hint": "Reserva vs estimativa fiscal"},
        {"key": "employees", "label": "Funcionários", "value": int(company.get("employees_count", 0)), "unit": "",
         "status": rag(m_income - payroll, 0.01, -1) if payroll else "green", "hint": "Custo de equipa sustentável"},
        {"key": "bank", "label": "Banco", "value": round(bank, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(bank, monthly_burn * 3, 0), "hint": "Saldo bancário estimado"},
        {"key": "risk", "label": "Risco", "value": round(runway, 1), "unit": "meses",
         "status": rag(runway, 6, 3), "hint": "Meses de autonomia de caixa"},
    ]
    status_score = {"green": 100, "amber": 55, "red": 20}
    health = round(sum(status_score[v["status"]] for v in vitals) / len(vitals))
    annual_profit = net if net > 0 else 0
    company_value = round(bank + annual_profit * 3, 2)
    dna = await db.ceo_dna.find_one({"user_id": user_id}) or {}
    goal_value = float(dna.get("target_revenue", 0)) or 1000000
    progress = min(100, round(company_value / goal_value * 100)) if goal_value else 0
    equity_progress = min(100, round(bal["net_worth"] / goal_value * 100)) if goal_value and bal["net_worth"] > 0 else 0

    return {
        "health": health, "vitals": vitals, "currency": currency,
        "currency_symbol": CURRENCY_SYMBOL.get(currency, "€"),
        "company_name": company.get("name", "A minha empresa"),
        "company_value": company_value, "goal_value": goal_value, "progress": progress,
        "cash_balance": round(bank, 2), "monthly_net": round(m_net, 2),
        "monthly_income": round(m_income, 2), "monthly_expense": round(m_expense, 2),
        "runway": round(runway, 1), "profit_margin": round(profit_margin, 1),
        "total_income": round(income, 2), "total_expense": round(expense, 2),
        "cash_available": bal["cash"], "total_assets": bal["total_assets"],
        "total_liabilities": bal["total_liabilities"], "net_worth": bal["net_worth"],
        "has_balance": bool(profile), "equity_progress": equity_progress,
    }

MODE_PROMPTS = {
    "conservador": "És prudente e avesso ao risco. Priorizas estabilidade, reservas de caixa e evitas dívida.",
    "crescimento": "És focado em crescimento sustentável. Equilibras oportunidade e risco.",
    "agressivo": "És ambicioso e orientado a resultados rápidos. Aceitas mais risco por retorno maior.",
    "familiar": "És equilibrado, valorizas qualidade de vida, tempo com a família e sustentabilidade do negócio.",
    "startup": "És orientado a escala, produto e captação. Pensas em métricas de crescimento e runway.",
    "investidor": "Pensas como investidor: retorno sobre capital, valor da empresa e saída (exit).",
}

PROFILE_LABELS = {
    "activity": "O que a empresa faz",
    "years_active": "Anos de atividade",
    "location": "Localização",
    "business_model": "Como ganha dinheiro",
    "avg_price": "Preço médio do produto/serviço",
    "biggest_client_pct": "Peso do maior cliente nas vendas (%)",
    "client_recurrence": "Os clientes voltam a comprar",
    "founder_dependency": "A empresa funciona sem o dono",
    "debt": "Dívidas / empréstimos",
    "biggest_cost": "Maior custo mensal",
    "supplier_dependency": "Depende muito de um fornecedor",
    "seasonality": "Meses fortes ou fracos",
    "cae": "CAE (código de atividade)",
    "main_goal": "Objetivo com a empresa",
    "personal_goal": "Objetivo pessoal do dono",
    "advantage": "O que a distingue da concorrência",
    "main_worry": "Maior preocupação atual",
}

async def build_system_prompt(user_id: str, user_name: str):
    settings = await db.settings.find_one({"user_id": user_id}) or {}
    mode = settings.get("ceo_mode", "crescimento")
    tone = settings.get("briefing_tone", "direto")
    dna = await db.ceo_dna.find_one({"user_id": user_id}) or {}
    memories = await db.memories.find({"user_id": user_id}).to_list(100)
    snap = await build_snapshot(user_id)
    company = await resolve_company(user_id)
    prof = (company or {}).get("profile", {}) or {}
    sector = (company or {}).get("sector") or prof.get("activity") or ""
    cae = prof.get("cae")
    sector_line = (f"Esta empresa opera no setor: {sector}" + (f" (CAE {cae})" if cae else "") + ".") if sector else "O setor da empresa ainda NÃO está indicado."
    prof_txt = "\n".join(f"{lbl}: {prof.get(k)}" for k, lbl in PROFILE_LABELS.items() if prof.get(k) not in (None, "", 0)) or "(o empresário ainda não preencheu o perfil da empresa)"
    mem_txt = "\n".join(f"- {m['content']}" for m in memories) or "- (ainda sem memórias registadas)"
    vitals_txt = "\n".join(f"- {v['label']}: {v['value']}{v['unit']} [{v['status']}]" for v in snap["vitals"])
    return (
        f"És o CEO AI — o Diretor Executivo Digital de {user_name}. NÃO és um chatbot nem um assistente técnico: "
        f"és um CEO experiente que já geriu centenas de empresas e que agora toma decisões LADO A LADO com este empresário. "
        f"A tua personalidade é experiente, calma, objectiva e confiante. {MODE_PROMPTS.get(mode, MODE_PROMPTS['crescimento'])} Tom: {tone}.\n\n"
        f"### COMO RESPONDES (obrigatório)\n"
        f"NUNCA respondas apenas com teoria e NUNCA digas 'depende'. Respondes sempre como um consultor executivo de topo, "
        f"tomando posição. Estrutura natural de cada resposta (sem cabeçalhos rígidos, de forma fluida e humana):\n"
        f"1) O QUE EU FARIA — a decisão concreta, na primeira pessoa e directa.\n"
        f"2) PORQUÊ — o raciocínio ligado aos números reais e aos objectivos pessoais do empresário.\n"
        f"3) RISCOS — o que pode correr mal.\n"
        f"4) ALTERNATIVAS — 1 ou 2 caminhos possíveis.\n"
        f"Foca-te no FUTURO e nas decisões, não no passado. Sê conciso, calmo e confiante. Fala português europeu.\n\n"
        f"### ESPECIALIZAÇÃO NO SETOR (OBRIGATÓRIO — nunca generalizes)\n"
        f"{sector_line}\n"
        f"Age como um CEO que conhece PROFUNDAMENTE este setor específico. Todos os conselhos, "
        f"referências (margens típicas, ticket médio, custos-chave, sazonalidade), riscos, KPIs e boas práticas "
        f"DEVEM ser próprios deste setor — usa o vocabulário e a realidade de quem gere este tipo de negócio "
        f"(ex.: uma construtora fala de obras, adjudicações, mão-de-obra e materiais; um restaurante fala de "
        f"food cost, rotação de mesas, ementa e turnos). Compara sempre com as referências típicas DESTE setor e "
        f"evita conselhos genéricos que serviriam para qualquer empresa. Se o setor não estiver indicado, "
        f"recomenda ao empresário preenchê-lo na área Empresa.\n\n"
        f"### PERFIL (CEO DNA)\n"
        f"Sonho: {dna.get('dream', 'n/d')}\nFaturação desejada: {dna.get('target_revenue', 'n/d')}\n"
        f"Horas de trabalho: {dna.get('work_hours', 'n/d')}\nPlano de saída: {dna.get('exit_plan', 'n/d')}\n"
        f"Visão a 5 anos: {dna.get('five_year_vision', 'n/d')}\n\n"
        f"### MEMÓRIA (lembra-te disto sempre)\n{mem_txt}\n\n"
        f"### PERFIL DA EMPRESA (informação dada pelo empresário — usa-a sempre na tua análise)\n{prof_txt}\n\n"
        f"### ESTADO ATUAL DA EMPRESA ({snap['company_name']})\n"
        f"Saúde: {snap['health']}/100\nCaixa: {snap['currency_symbol']}{snap['cash_balance']}\n"
        f"Resultado mensal: {snap['currency_symbol']}{snap['monthly_net']}\nAutonomia: {snap['runway']} meses\n"
        f"Valor da empresa: {snap['currency_symbol']}{snap['company_value']} (objetivo {snap['currency_symbol']}{snap['goal_value']}, {snap['progress']}%)\n"
        f"Sinais vitais:\n{vitals_txt}"
    )

async def get_chat(user_id: str, user_name: str, session_id: str):
    settings = await db.settings.find_one({"user_id": user_id}) or {}
    provider, model = MODEL_MAP.get(settings.get("model", "claude"), MODEL_MAP["claude"])
    sysmsg = await build_system_prompt(user_id, user_name)
    if provider == "anthropic":
        chat = LlmChat(api_key=EMERGENT_KEY, session_id=session_id, system_message=sysmsg,
                       custom_headers={"anthropic-beta": "task-budgets-2026-03-13"}).with_model(provider, model)
        chat = chat.with_params(extra_body={"output_config": {"task_budget": {"type": "tokens", "total": 200000}, "effort": "high"}}, max_tokens=8000)
    else:
        chat = LlmChat(api_key=EMERGENT_KEY, session_id=session_id, system_message=sysmsg).with_model(provider, model)
    return chat

async def make_briefing(user_id: str, user_name: str):
    settings = await db.settings.find_one({"user_id": user_id}) or {}
    count = settings.get("briefing_count", 4)
    snap = await build_snapshot(user_id)
    sysmsg = await build_system_prompt(user_id, user_name)
    hour = datetime.now(timezone.utc).hour
    greeting = "Bom dia" if hour < 12 else ("Boa tarde" if hour < 19 else "Boa noite")
    prompt = (
        f"Gera o briefing diário para o empresário. Devolve APENAS JSON válido no formato: "
        f'{{"greeting":"{greeting}, <nome>. ...","items":[{{"title":str,"detail":str,"priority":"alta"|"media"|"baixa","icon":"cash"|"profit"|"clients"|"tax"|"risk"|"opportunity"}}]}}. '
        f"Exatamente {count} itens, priorizados pelo que mais importa hoje. "
        f"O greeting deve ser uma frase humana e calorosa a começar com '{greeting}'. Detalhes curtos, orientados ao futuro e à ação. Sem texto fora do JSON."
    )
    chat = LlmChat(api_key=EMERGENT_KEY, session_id=f"brief-{uuid.uuid4()}", system_message=sysmsg).with_model("openai", "gpt-5.4")
    try:
        resp = await chat.send_message(UserMessage(text=prompt))
        text = resp.strip()
        if "```" in text:
            text = text.split("```")[1].replace("json", "", 1).strip()
        data = json.loads(text)
    except Exception as e:
        logger.error(f"briefing error: {e}")
        data = {"greeting": f"{greeting}, {user_name}. Aqui está o que precisa da sua atenção hoje.",
                "items": [{"title": "Ligue os seus dados", "detail": "Registe receitas e despesas para eu analisar a saúde da sua empresa.",
                           "priority": "alta", "icon": "opportunity"}]}
    data["health"] = snap["health"]
    return data

PRIORITY_COLOR = {"alta": "#EF4444", "media": "#F59E0B", "baixa": "#10B981"}

def build_briefing_html(name: str, data: dict, app_url: str):
    rows = ""
    for it in data.get("items", []):
        pc = PRIORITY_COLOR.get(it.get("priority", "media"), "#F59E0B")
        rows += f"""
        <tr><td style="padding:0 0 14px 0;">
          <table width="100%" cellpadding="0" cellspacing="0" style="background:#faf9f6;border:1px solid #eee;border-radius:12px;">
            <tr>
              <td width="6" style="background:{pc};border-radius:12px 0 0 12px;">&nbsp;</td>
              <td style="padding:14px 18px;">
                <div style="font-size:15px;font-weight:700;color:#18181b;">{it.get('title','')}</div>
                <div style="font-size:14px;color:#52525b;margin-top:4px;line-height:1.5;">{it.get('detail','')}</div>
              </td>
            </tr>
          </table>
        </td></tr>"""
    return f"""<!DOCTYPE html><html><body style="margin:0;background:#0b0c10;font-family:Arial,Helvetica,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#0b0c10;padding:32px 0;">
      <tr><td align="center">
        <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:18px;overflow:hidden;">
          <tr><td style="background:#0b0c10;padding:28px 32px;">
            <div style="color:#D4AF37;font-size:22px;font-weight:700;letter-spacing:1px;">CEO&nbsp;AI</div>
            <div style="color:#a1a1aa;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Executivo Digital · Briefing Diário</div>
          </td></tr>
          <tr><td style="padding:32px;">
            <div style="font-size:22px;color:#18181b;font-weight:700;line-height:1.35;margin-bottom:8px;">{data.get('greeting','Bom dia')}</div>
            <div style="font-size:13px;color:#71717a;margin-bottom:22px;">Saúde da empresa: <strong style="color:#D4AF37;">{data.get('health',0)}/100</strong></div>
            <table width="100%" cellpadding="0" cellspacing="0">{rows}</table>
            <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:22px;"><tr><td align="center">
              <a href="{app_url}" style="display:inline-block;background:#D4AF37;color:#0b0c10;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;">Abrir o meu CEO AI</a>
            </td></tr></table>
          </td></tr>
          <tr><td style="padding:20px 32px;background:#faf9f6;border-top:1px solid #eee;">
            <div style="font-size:11px;color:#a1a1aa;">Recebes este email porque ativaste o briefing diário. Podes desativar em Personalização.</div>
          </td></tr>
        </table>
      </td></tr>
    </table></body></html>"""

async def send_email_raw(to_email: str, subject: str, html: str):
    if not EMAIL_KEY:
        logger.error("EMERGENT_EMAIL_KEY not set")
        return False
    payload = {"to": [to_email], "subject": subject, "html": html, "from_name": EMAIL_FROM_NAME}
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                                     headers={"X-Email-Key": EMAIL_KEY}, json=payload)
        resp.raise_for_status()
        return True
    except Exception as e:
        logger.error(f"email send error: {e}")
        return False

# ---------------------------------------------------------------- password reset
def hash_reset_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()

async def create_password_reset(user_id: str) -> str:
    raw = secrets.token_urlsafe(32)
    await db.password_reset_tokens.insert_one({
        "user_id": user_id,
        "token_hash": hash_reset_token(raw),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
    })
    return raw

def build_reset_password_html(name: str, link: str) -> str:
    who = f", {name}" if name else ""
    return (f"<div style='font-family:Arial,sans-serif;max-width:560px;margin:auto;padding:24px'>"
            f"<h2 style='color:#0b0c10'>Redefinicao de senha - CEO AI</h2>"
            f"<p>Ola{who}, recebemos um pedido para redefinir a senha da tua conta CEO AI.</p>"
            f"<p style='margin:24px 0'><a href='{link}' style='display:inline-block;background:#3B82F6;color:#fff;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;'>Definir nova senha</a></p>"
            f"<p style='font-size:13px;color:#71717a'>Esta ligacao e valida por 1 hora e so pode ser usada uma vez. Se nao pediste esta alteracao, ignora este email.</p>"
            f"<p style='font-size:12px;color:#a1a1aa;word-break:break-all'>{link}</p>"
            f"</div>")

async def send_password_reset_email(user_doc: dict) -> bool:
    raw = await create_password_reset(str(user_doc["_id"]))
    frontend = os.environ.get("FRONTEND_URL", "").rstrip("/")
    link = f"{frontend}/reset-password?token={raw}"
    html = build_reset_password_html(user_doc.get("name", ""), link)
    return await send_email_raw(user_doc.get("email", ""), "Redefinicao de senha - CEO AI", html)


async def send_daily_briefings():
    today = datetime.now(timezone.utc).date().isoformat()
    cursor = db.settings.find({"email_briefing": True})
    async for s in cursor:
        uid = s.get("user_id")
        if not uid:
            continue
        claim = await db.settings.update_one(
            {"user_id": uid, "email_briefing": True, "last_briefing_email_date": {"$ne": today}},
            {"$set": {"last_briefing_email_date": today}})
        if claim.modified_count != 1:
            continue
        try:
            u = await db.users.find_one({"_id": ObjectId(uid)})
            if not u:
                continue
            data = await make_briefing(uid, u.get("name", ""))
            html = build_briefing_html(u.get("name", ""), data, os.environ.get("FRONTEND_URL", ""))
            await send_email_raw(u["email"], "O teu briefing diário — CEO AI", html)
        except Exception as e:
            logger.error(f"daily briefing error for {uid}: {e}")

async def ai_json(system: str, prompt: str, model=("openai", "gpt-5.4")):
    chat = LlmChat(api_key=EMERGENT_KEY, session_id=f"j-{uuid.uuid4()}", system_message=system).with_model(*model)
    try:
        resp = await chat.send_message(UserMessage(text=prompt))
        t = resp.strip()
        if "```" in t:
            t = t.split("```")[1].replace("json", "", 1).strip()
        return json.loads(t)
    except Exception as e:
        logger.error(f"ai_json error: {e}")
        return None

async def cached_ai(kind: str, uid: str, cid, system: str, prompt: str):
    today = datetime.now(timezone.utc).date().isoformat()
    q = {"kind": kind, "user_id": uid, "company_id": cid, "date": today}
    hit = await db.ai_cache.find_one(q)
    if hit and hit.get("payload"):
        return hit["payload"]
    payload = await ai_json(system, prompt)
    if payload:
        await db.ai_cache.update_one(q, {"$set": {"payload": payload, "created_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return payload

async def invalidate_ai_cache(uid: str):
    await db.ai_cache.delete_many({"user_id": uid})

async def _growth_score(uid: str, cid: str):
    entries = await db.entries.find({"user_id": uid, "company_id": cid}, {"type": 1, "amount": 1, "date": 1}).to_list(5000) if cid else []
    inc = {}
    for e in entries:
        mk = str(e.get("date", ""))[:7]
        if len(mk) == 7 and e["type"] == "income":
            inc[mk] = inc.get(mk, 0) + e["amount"]
    sm = sorted(inc)
    g = 50
    if len(sm) >= 2:
        recent = sum(inc[m] for m in sm[-3:]); prior = sum(inc[m] for m in sm[-6:-3])
        if prior > 0:
            g = max(5, min(100, int(60 + ((recent - prior) / prior) * 100)))
        elif recent > 0:
            g = 72
    return g

# ================================================================ FOUNDER CAMPAIGN / BILLING / ADMIN
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError

FOUNDER_LIMIT = 15
FOUNDER_PRICE_MONTHLY = 29
PROFESSIONAL_PRICE_MONTHLY = 59
ENTERPRISE_PRICE_MONTHLY = 159.99
PROFESSIONAL_TRIAL_DAYS = 7
FOUNDER_PROGRAM_ACTIVE_DEFAULT = True
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "").lower()

LOOKUP_TO_PLAN = {
    "founder_monthly": "founder",
    "professional_monthly": "professional",
    "premium_monthly": "professional",
    "premium_yearly": "professional",
}
PLAN_LABELS = {"founder": "Empresa Fundadora", "professional": "Professional", "enterprise": "Enterprise"}
PLAN_PRICE = {"founder": FOUNDER_PRICE_MONTHLY, "professional": PROFESSIONAL_PRICE_MONTHLY, "enterprise": ENTERPRISE_PRICE_MONTHLY}
PREMIUM_STATUSES = {"active", "trialing"}

def plan_from_lookup(lk):
    return LOOKUP_TO_PLAN.get(lk, "professional")

def is_admin_email(user: dict) -> bool:
    return bool(ADMIN_EMAIL) and (user.get("email", "") or "").lower() == ADMIN_EMAIL

async def get_admin_user(user: dict = Depends(get_current_user)) -> dict:
    if not is_admin_email(user):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador")
    return user

async def can_access_premium(user: dict) -> bool:
    if is_admin_email(user):
        return True
    return bool(user.get("is_premium"))

async def premium_user(user: dict = Depends(get_current_user)) -> dict:
    if not (is_admin_email(user) or bool(user.get("is_premium"))):
        raise HTTPException(status_code=402, detail="premium_required")
    return user

# ---------------------------------------------------------------- campaign config
async def get_campaign() -> dict:
    c = await db.app_config.find_one({"_id": "founder_campaign"})
    if not c:
        c = {"_id": "founder_campaign", "active": FOUNDER_PROGRAM_ACTIVE_DEFAULT, "milestones_sent": []}
        try:
            await db.app_config.insert_one(c)
        except DuplicateKeyError:
            c = await db.app_config.find_one({"_id": "founder_campaign"})
    return c

async def set_campaign_active(value: bool):
    await db.app_config.update_one({"_id": "founder_campaign"}, {"$set": {"active": bool(value)}}, upsert=True)

async def founder_claimed_count() -> int:
    doc = await db.counters.find_one({"_id": "founder"})
    return int((doc or {}).get("seq", 0))

async def _allocate_founder_number():
    doc = await db.counters.find_one_and_update(
        {"_id": "founder", "seq": {"$lt": FOUNDER_LIMIT}},
        {"$inc": {"seq": 1}},
        return_document=ReturnDocument.AFTER,
    )
    return int(doc["seq"]) if doc else None

async def handle_founder_activation(user_doc: dict):
    """Atomic, race-safe founder slot allocation. Returns founder_number or None."""
    if not user_doc:
        return None
    oid = user_doc["_id"]
    if user_doc.get("founder_number") or user_doc.get("is_founder"):
        return None  # already a founder (historical) — never reallocate
    camp = await get_campaign()
    if not camp.get("active", True):
        return None
    # per-user lock: only one concurrent activation can claim
    claim = await db.users.update_one(
        {"_id": oid, "founder_number": {"$exists": False}, "is_founder": {"$ne": True},
         "founder_claim_in_progress": {"$ne": True}},
        {"$set": {"founder_claim_in_progress": True}})
    if claim.modified_count != 1:
        return None
    num = await _allocate_founder_number()
    if not num:
        await db.users.update_one({"_id": oid}, {"$unset": {"founder_claim_in_progress": ""}})
        await set_campaign_active(False)
        return None
    now = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"_id": oid}, {
        "$set": {"is_founder": True, "founder_number": num, "founder_activated_at": now,
                 "founder_price_locked": True, "founder_subscription_status": "active"},
        "$unset": {"founder_claim_in_progress": ""}})
    remaining = FOUNDER_LIMIT - num
    try:
        await notify_founder_activated(user_doc, num, remaining)
        await check_founder_milestones(remaining)
    except Exception as e:
        logger.error(f"founder notify error: {e}")
    if num >= FOUNDER_LIMIT:
        await set_campaign_active(False)
    return num

# ---------------------------------------------------------------- stripe subscription sync
def _sub_period_end(sub, item):
    return sub.get("current_period_end") or (item.get("current_period_end") if item else None)

async def sync_subscription(sub_id: str, user_id: str = None):
    if not sub_id:
        return
    try:
        sub = stripe.Subscription.retrieve(sub_id, expand=["items.data.price"])
    except Exception as e:
        logger.error(f"sync_subscription retrieve error: {e}")
        return
    items = sub.get("items", {}).get("data", [])
    item = items[0] if items else None
    price = item["price"] if item else {}
    lk = price.get("lookup_key")
    status = sub.get("status")
    cpe = _sub_period_end(sub, item)
    customer = sub.get("customer")
    md = sub.get("metadata") or {}
    uid = user_id or md.get("user_id")
    user_doc = None
    if uid:
        try:
            user_doc = await db.users.find_one({"_id": ObjectId(uid)})
        except Exception:
            user_doc = None
    if not user_doc and customer:
        user_doc = await db.users.find_one({"stripe_customer_id": customer})
    if not user_doc and sub_id:
        user_doc = await db.users.find_one({"stripe_subscription_id": sub_id})
    if not user_doc:
        logger.error(f"sync_subscription: no user for {sub_id}")
        return
    plan = plan_from_lookup(lk)
    premium = status in PREMIUM_STATUSES
    upd = {"stripe_customer_id": customer, "stripe_subscription_id": sub_id,
           "subscription_status": status, "plan": plan, "is_premium": premium,
           "current_period_end": cpe, "subscription_lookup_key": lk,
           "cancel_at_period_end": bool(sub.get("cancel_at_period_end"))}
    if premium and not user_doc.get("subscription_started_at"):
        upd["subscription_started_at"] = datetime.now(timezone.utc).isoformat()
    if status in ("canceled", "unpaid", "incomplete_expired"):
        upd["is_premium"] = False
        upd["subscription_cancelled_at"] = datetime.now(timezone.utc).isoformat()
        if user_doc.get("is_founder"):
            upd["founder_price_locked"] = False
            upd["founder_subscription_status"] = "cancelled"
    await db.users.update_one({"_id": user_doc["_id"]}, {"$set": upd})
    if plan == "founder" and status == "active":
        fresh = await db.users.find_one({"_id": user_doc["_id"]})
        await handle_founder_activation(fresh)

# ---------------------------------------------------------------- admin notifications
async def notify_founder_activated(user_doc: dict, num: int, remaining: int):
    company = await resolve_company(str(user_doc["_id"]))
    cname = (company or {}).get("name", "(empresa)")
    name = user_doc.get("name", ""); email = user_doc.get("email", "")
    now = datetime.now(timezone.utc)
    await db.admin_notifications.insert_one({
        "type": "founder_activated", "founder_number": num, "company": cname,
        "name": name, "email": email, "remaining": remaining, "read": False,
        "created_at": now.isoformat()})
    subject = f"Nova Empresa Fundadora ativada — vaga {num} de {FOUNDER_LIMIT}"
    html = (f"<div style='font-family:Arial,sans-serif;max-width:560px;margin:auto'>"
            f"<h2 style='color:#0b0c10'>Nova Empresa Fundadora ativada</h2>"
            f"<p>Uma nova Empresa Fundadora concluiu a subscrição.</p>"
            f"<table cellpadding='6' style='font-size:14px'>"
            f"<tr><td><b>Empresa</b></td><td>{cname}</td></tr>"
            f"<tr><td><b>Responsável</b></td><td>{name}</td></tr>"
            f"<tr><td><b>E-mail</b></td><td>{email}</td></tr>"
            f"<tr><td><b>Posição</b></td><td>{num} de {FOUNDER_LIMIT}</td></tr>"
            f"<tr><td><b>Preço</b></td><td>{FOUNDER_PRICE_MONTHLY} €/mês</td></tr>"
            f"<tr><td><b>Data e hora</b></td><td>{now.strftime('%d/%m/%Y %H:%M UTC')}</td></tr>"
            f"<tr><td><b>Vagas restantes</b></td><td>{remaining}</td></tr>"
            f"</table></div>")
    if ADMIN_EMAIL:
        await send_email_raw(ADMIN_EMAIL, subject, html)

async def check_founder_milestones(remaining: int):
    if remaining not in (5, 3, 1, 0):
        return
    camp = await get_campaign()
    sent = camp.get("milestones_sent", []) or []
    if remaining in sent:
        return
    await db.app_config.update_one({"_id": "founder_campaign"}, {"$addToSet": {"milestones_sent": remaining}}, upsert=True)
    if remaining == 0:
        subject = f"Programa Empresas Fundadoras concluído — {FOUNDER_LIMIT} de {FOUNDER_LIMIT} vagas preenchidas."
        body = "Todas as vagas de Empresa Fundadora foram preenchidas. O plano Professional continua disponível."
    else:
        subject = f"Programa Empresas Fundadoras — {'resta' if remaining == 1 else 'restam'} {remaining} {'vaga' if remaining == 1 else 'vagas'}"
        body = f"Restam apenas {remaining} vagas de Empresa Fundadora."
    await db.admin_notifications.insert_one({"type": "milestone", "remaining": remaining,
                                             "read": False, "created_at": datetime.now(timezone.utc).isoformat(), "text": subject})
    if ADMIN_EMAIL:
        html = f"<div style='font-family:Arial,sans-serif'><h2>{subject}</h2><p>{body}</p></div>"
        await send_email_raw(ADMIN_EMAIL, subject, html)

# ---------------------------------------------------------------- audit
async def audit_log(admin_email: str, action: str, target: str = None, before=None, after=None):
    await db.audit_log.insert_one({"admin": admin_email, "action": action, "target": target,
                                   "before": before, "after": after,
                                   "created_at": datetime.now(timezone.utc).isoformat()})
